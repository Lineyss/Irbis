from django.contrib.auth.forms import UserCreationForm
from .models_fields import get_AMains_by_id_component
from django.views.generic import CreateView
from django.urls import reverse_lazy
from openpyxl import load_workbook
from typing import Any, Type
from django import forms
from .models import *

def get_model_form(Model:Type[models.Model]):
    """
    Метод генерации формы на основе переданной модели
    """
    if Model._meta.concrete_model is PCB:
        class PCBForm(forms.ModelForm):
            gerber_load = forms.FileField(label='Гербер файлы', widget=forms.ClearableFileInput(attrs={'allow_multiple_selected': True}), required=False)

            class Meta:
                model = Model
                fields = '__all__'
            
            def __init__(self, *args, **kwargs):
                self.instance = kwargs.get('instance', None)
                super(PCBForm, self).__init__(*args, **kwargs)

            def clean(self) -> dict[str, Any]:
                clean_data = super().clean()
                gerber_load = clean_data.get('gerber_load')

                if gerber_load is not None:
                    try:
                        self.gerber = Gerber.objects.create(file=gerber_load)
                        self.gerber.save()
                    except Exception as e:
                        print(e)
                        raise forms.ValidationError('Не верный формат гербер файла')
                
                return clean_data
                
            def save(self, commit=True):
                pcb = super().save(commit=commit)
                pcb.gerber_files.add(self.gerber)
                pcb.save()
                return pcb

        return PCBForm
    else:
        class FurnitureFrom(forms.ModelForm):
            class Meta:
                model = Model
                fields = '__all__'

        return FurnitureFrom

def get_createViewModel(Model:Type[models.Model]):
    """
    Метод генерации CreateView
    """

    if Model._meta.concrete_model is PCB:
        class FurnitureViewModel(CreateView):
            model = Model
            form_class  = get_model_form(Model)
            template_name = 'main/category_view.html'
            success_url = reverse_lazy('main_index')
            
            def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
                data = super().get_context_data(**kwargs)
                data['title'] = Model._meta.verbose_name

                return data
                
        return FurnitureViewModel.as_view()
    else:
        class FurnitureViewModel(CreateView):
            model = Model
            fields = '__all__'
            template_name = 'main/category_view.html'
            success_url = reverse_lazy('main_index')
            
            def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
                data = super().get_context_data(**kwargs)
                data['title'] = Model._meta.verbose_name

                return data
                
        return FurnitureViewModel.as_view()

class RegisterView(CreateView):
    form_class = UserCreationForm
    success_url = reverse_lazy('login')
    template_name = "registration/reg.html"
    redirect_authenticated_user = True

class PCB_Upload_FileForm(forms.ModelForm):

    def clean(self) -> dict[str, Any]:
        cleaned_data = super().clean()

        print(cleaned_data)

        list_components = cleaned_data.get('list_components')
        electrical_diagram = cleaned_data.get('electrical_diagram')
        assembly_drawing = cleaned_data.get('assembly_drawing')

        isGood = False
        errors = []

        if list_components is not None:
            if type(list_components) != bool:
                if list_components.name.endswith(".xls") == False:
                    errors.append("Перечень компонентов может быть только в формате .xls")
                elif self.check_list_components(list_components) == False:
                    errors.append("Таблице не валидная")
                else:
                    isGood = True
            else:
                isGood = True

        if  electrical_diagram is not None:
            if type(electrical_diagram) != bool and electrical_diagram.name.endswith(".pdf") == False:
                errors.append("Электрическая схема может быть только в формате .pdf")
            else:
                isGood = True
            
        if assembly_drawing is not None:
            if type(electrical_diagram) != bool and assembly_drawing.name.endswith(".pdf") == False:
                errors.append("Сборный чертеж может быть только в формате .pdf")
            else:
                isGood = True

        if errors and isGood:
            for error in errors:
                raise forms.ValidationError(error)


        return cleaned_data

    def check_list_components(self, file)->bool:
        wb = load_workbook(file)
        worksheet = wb.active
        return True

    def check_rows(self, worksheet, col, Type) -> bool:
        for row in range(1, worksheet.max_row ):
            if not isinstance(col[row].value, Type) and col[row].value != '-':
                return False
        return True
    
    class Meta:
        model = PCB
        fields = ['list_components', 'electrical_diagram', 'assembly_drawing']
