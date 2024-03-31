# from django.http import HttpResponse
from .models_fields import get_AMains_by_id_component,get_furnitures
from openpyxl import load_workbook
from django.http import *
from Irbis.settings import BASE_DIR
from .models import *
from PIL import Image
from .forms import *
from pygerber.gerberx3.api import (
      ColorScheme,
      Rasterized2DLayer,
      Rasterized2DLayerParams,
)
import os


def module_detail_view(model)-> dict:
    data = get_data()

    moduleCompositions = ModuleComposition.objects.filter(module = model)

    for moduleComposition in moduleCompositions:

        elements = PCBComposition.objects.filter(pcb = moduleComposition.pcb)

        data = get_all_connected_element(elements, data)

    data['file'] = model.stepFile

    return data

def pcb_detail_view(model, gbrLoad)-> dict:
    data = get_data()
    data['gbrs'] = None
    data['buildFiles'] = None

    elements = PCBComposition.objects.filter(pcb = model)

    data = get_all_connected_element(elements, data)

    fields = data['element']['fields']

    if checkFileExist(model.list_components):
        add_list_components(model)   

    if fields is not None:
        count_delete = 0
        for i in range(len(fields)):
            i -= count_delete
            if fields[i]._verbose_name == 'Корпус компонента' or fields[i]._verbose_name == 'Категория':
                fields.pop(i)
                count_delete += 1

        data['element']['fields'].append(CustomField('Need', 'Нужно'))
        
    if data['element']['items'] is not None:
        for i in range(len(data['element']['items'])):
            element = data['element']['items'][i]
            element.Need = elements[i].need_quantity

    
    if gbrLoad == 'True':
        dict = upload_gerber_files(model)
        data['gbrs'] = dict['title_file']
        data['buildFiles'] = dict['png_path']

    return data

def search_model_by_pk(name, pk)->object:
    category = Category.objects.get(name = name)
    model = get_furnitures(category)['model']
    return model.objects.get(pk = pk)

def get_all_connected_element(elements, data = None)-> dict:
    data = get_data(data)

    for element in elements:
        if(element.element is None): continue

        AMain_item = get_AMains_by_id_component(element.element.pk)

        if AMain_item['object'] is None: continue

        if  data['element']['fields'] is None:
            data['element']['fields'] = AMain_item['fields']
            data['element']['fields'].append(CustomField('Category', 'Категория'))

            if type(AMain_item['object']) is Other:
                data['element']['fields'].append(CustomField('Package','Корпус компонента'))

        objects = data['element']['items']
        objects = list() if objects is None else list(objects)
        AMain_item['object'].Category = AMain_item['object'].category.name
        objects.append(AMain_item['object'])
        data['element']['items'] = objects

    return data

def get_data(data = None)-> dict:
    return data if data is not None else {
        'element': {
            'fields': None,
            'items': None
        },
    }

def check_valid_files(list_components,electrical_diagram, assembly_drawing)->bool:
    if list_components is None or electrical_diagram is None or assembly_drawing is None:
        return False
    
    return True

def check_list_components(IDC, module)->bool:
    try:
        pcb = PCBComposition.objects.get(pcb = module, element = IDC)
        return True
    except:
        return False

def find_number_column(title:str, cols):
    i = 0
    for col in cols:
        if col[0].value == title:
            return i
        i += 1

def add_list_components(model):
    wb = load_workbook(model.list_components)
    worksheet = wb.active

    i_IDC = find_number_column('Идентификатор компонента', worksheet.iter_cols())
    i_Need = find_number_column('Нужно', worksheet.iter_cols())
    i = 0

    for row in worksheet.iter_rows():
        if i == 0:
            i += 1
            continue

        element = UniqueComponent.objects.get(pk = row[i_IDC].value)
        need = row[i_Need].value

        if check_list_components(element, model) == False:
            pcbC = PCBComposition.objects.create(pcb = model, element = element, need_quantity = need)
            pcbC.save()

def checkFileExist(file)-> bool:
    try:
        file.url
        return True
    except:
        return False

def upload_components_files(request, model):
    form = PCB_Upload_FileForm(request.POST, request.FILES,  instance=model)
    if form.is_valid():
        form.save()
        return HttpResponse(status=200)
    
    return HttpResponseBadRequest(form.errors.values)

def upload_gerber_files(model)->dict:
    data = {
        'title_file': [],
        'png_path': [],
    }

    print(BASE_DIR)

    gerber = model.gerber_files.all()
    for file in gerber:
        path = str(file)
        path = path.split('/')
        name = path[len(path) - 1]

        data['title_file'].append(name)
        data['png_path'] = convert_gerber_to_png(str(file), name)

    return data 

def crop_image(path):
    image = Image.open(path)

    left = 0
    top = 0
    right = 1600
    bottom = 2100

    cropped_image = image.crop((left, top, right, bottom))

    cropped_image.save(path)

def resize(path):
    background_image = Image.new('RGBA', (1600, 2000), (255, 255, 255, 0))
    name = os.path.basename(path)
    base_dir = f"{BASE_DIR}\media\images\help\{name}.png"
    print(base_dir)
    background_image.save(base_dir)
    background_image = Image.open(base_dir)

    overlay_image = Image.open(path)
    bg_width, bg_height = background_image.size
    overlay_width, overlay_height = overlay_image.size
    position = ((bg_width - overlay_width) // 2, (bg_height - overlay_height) // 2)
    background_image.paste(overlay_image, position)
    background_image.save(path)

    # os.remove(base_dir)

def convert_gerber_to_png(file_path, file_name = None)->str:
    file_path = f"{BASE_DIR}\\media\\{file_path}"
    lines = ''
    color=''
    extends = os.path.basename(file_path).split('.')[1]
    
    if file_name is None:
        file_name = file_path.split('/')
        file_name = file_path[len(file_name) - 1]

    png_path = f"{BASE_DIR}\media\images\gerber\{file_name}.png"

    with open(file_path, 'r') as file:
        lines = file.readlines()

    with open(file_path, 'w') as file:
        for line in lines:
            if 'G04' not in line:
                file.write(line)
    
    try:
        if extends == 'GTL' or extends == 'GBL' or extends == 'GKO':
            color = ColorScheme.COPPER_ALPHA
        else:
            color = ColorScheme.SILK_ALPHA
        Rasterized2DLayer(  
            options=Rasterized2DLayerParams(
                source_path=file_path,
                colors = color,
            ),
        ).render().save(png_path)
        image = Image.open(png_path)
        original_width, original_height = image.size
        print('--------------')
        print(png_path)
        print(f'{original_width} - {original_height}')
        # if original_width < 1600 and original_height < 2100:
        #     print('resize')
        #     resize(png_path)
        # else:
        #     print('crop_image')
        crop_image(png_path)
        print('--------------')
    except Exception as e:
        print(str(e))
        return None

    return png_path